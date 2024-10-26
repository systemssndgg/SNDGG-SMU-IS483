import constants as constants 
from landtransportsg import Traffic
import requests
from openai import OpenAI
import openpyxl
import os
from collections import defaultdict

import json
from requests.exceptions import RequestException, HTTPError
from ngsildclient import Client, Entity, SmartDataModels
from datetime import datetime
import pandas as pd

client = OpenAI(
    api_key = constants.OPENAI_KEY
)

API_KEY = constants.LTA_API_KEY
ctx = constants.ctx
broker_url = constants.broker_url
broker_port = constants.broker_port  # default, 80
temporal_port = constants.temporal_port  # default 1026
broker_tenant = constants.broker_tenant

price_dictionary = {
                    "WeekdayStr": "-",
                    "SaturdayStr": "-",
                    "SundayPHStr": "-",
                    "Car" :{
                            "EntryFees" : {
                                "Weekdays" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                },
                                "Saturday" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                },
                                "SundayPH" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                }
                            },
                            "WeekdayRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "weekdayMin" : None,
                                "weekdayRate" : None
                            },
                            "SaturdayRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "satdayMin" : None,
                                "satdayRate" : None
                            },
                            "SundayPHRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "sunPHMin" : None,
                                "sunPHRate" : None
                            },
                    },
                    "Motorcycle" :{
                            "EntryFees" : {
                                "Weekdays" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                },
                                "Saturday" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                },
                                "SundayPH" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                }
                            },
                            "WeekdayRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "weekdayMin" : None,
                                "weekdayRate" : None
                            },
                            "SaturdayRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "satdayMin" : None,
                                "satdayRate" : None
                            },
                            "SundayPHRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "sunPHMin" : None,
                                "sunPHRate" : None
                            },
                    },
                    "Heavy Vehicle" :{
                            "EntryFees" : {
                                "Weekdays" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                },
                                "Saturday" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                },
                                "SundayPH" : {
                                    "startTime" : None,
                                    "endTime" : None,
                                    "entryFee" : None
                                }
                            },
                            "WeekdayRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "weekdayMin" : None,
                                "weekdayRate" : None
                            },
                            "SaturdayRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "satdayMin" : None,
                                "satdayRate" : None
                            },
                            "SundayPHRate" : {
                                "startTime" : None,
                                "endTime" : None,
                                "sunPHMin" : None,
                                "sunPHRate" : None
                            },
                    }
                }

# (1) Fetch carpark availability from the data.gov.sg API

def fetch_carpark_availabilities():
    '''
    Inputs: None
    
    \nOutput: entity_dict (dict) A dictionary of carpark entities with the CarParkID as the key

    \nDescription: Fetches the 40 carpark availabilities under LTA and returns a dictionary of carpark entities with the CarParkID as the key.
    '''
    url = "https://datamall2.mytransport.sg/ltaodataservice/CarParkAvailabilityv2"
    commerical_ids = []

    # make a request to the url and retrieve the response
    response = requests.get(url, headers={
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0",
        "AccountKey": "iGUugKJzSkGPCPBbLAhyhA=="
    })
    
    # check if the response is successful
    if response.status_code != 200:
        raise Exception(f"Failed to fetch carpark availabilities: {response.status_code}")
    else:
        print("Successfully fetched carpark availabilities")
        response = response.json()
    
        # the response returns a JSON object with 'value' as a key, this in turns points to the list of carpark dictionaries.
        # iterate through the list of dictionaries
        entity_dict = {}
        for i in range(0, 40):
            entity = response['value'][i]
            entity_dict[entity['CarParkID']] = entity
        
        return entity_dict
    
# (2) Check if entities already exist in the broker

def create_entities_in_broker(entities):
    with Client(
        hostname=broker_url,
        port=broker_port,
        tenant=broker_tenant,
        port_temporal=temporal_port,
    ) as client:
        count = 0
        for entity in entities:
            ret = client.upsert(entity)
            if ret:
                count += 1
    print("Uploaded ", count)
    return ret

def check_entity_exists(entity):
    with Client(
        hostname=broker_url,
        port=broker_port,
        tenant=broker_tenant,
        port_temporal=temporal_port,
    ) as client:
        ret = client.query(entity)
        # iterate through the list of entities returned
        for i in range(len(ret)):
            # match the entity to the name of the carpark
            count = 1
    return ret

# (3) Fetch the carpark rates from data.gov if needed
def fetch_carpark_rates():
    '''
    Inputs -> None
    Outputs -> entity_dict : dict : A dictionary of carpark entities with the CarParkID as the key
    '''

    datasetId = "d_9f6056bdb6b1dfba57f063593e4f34ae"
    url = "https://data.gov.sg/api/action/datastore_search?resource_id="  + datasetId

    # make a request to the url and retrieve the response
    response = requests.get(url)
    
    # check if the response is successful
    if response.status_code != 200:
        raise Exception(f"Failed to fetch carpark rates: {response.status_code}")
    else:
        print("Successfully fetched carpark rates")
        response = response.json()
    
        # the response returns a JSON object with 'value' as a key, this in turns points to the list of carpark dictionaries.
        # iterate through the list of dictionaries
        return response['result']['records']

# (4) Use LLM to format the data into a dict template
def read_excel(file_path):
    '''
    Inputs: file_path : str : The path to the excel file to read
    \nOutputs: data : list : A list of lists, where each list corresponds to a row, containing the data from the excel file

    \nDescription: Reads the excel file at the given file path and returns the data in the form of a list of lists.
    '''
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = [row for row in sheet.iter_rows(values_only=True)]
    return data

def ask_gpt_json(system_prompt, user_prompt):
        '''
            Input:
                system_prompt: str : The system prompt to provide to the model 
                user_prompt: str : The user prompt to provide to the model 

            Output:
                response: dict : The response generated by the model 
                dict structure based on the examples given in system prompt.
                cost: float : The cost of the API call
        '''

        model="gpt-4.o",
        completion = client.chat.completions.create(
            # Adjust the temperature to control the randomness of the output
            temperature=0.2,
            model="gpt-4o-mini",
            # Output will be a string in valid JSON format, so we can parse it with json.loads
            response_format= {"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": system_prompt
                },
                {
                    "role": "user",
                    "content": user_prompt
                }
            ]
        )
        response = completion.choices[0].message.content
        try:
            f = open("response.txt", "w")
            f.write(response)
            f.close()
            print("Data formatted and written to file")
        except:
                print("Error writing data to file")

        # Converts the response to a JSON Object / Dictionary.
        response = json.loads(response)
        return response

def format_carpark_rates(carpark_rates):
    schema = '''
{
  "Carpark A": {
    "rates": {
      "weekday": {
        "time_based": [
          {
            "start_time": "07:30",
            "end_time": "17:00",
            "rate_per_hour": 2.0
          },
          {
            "start_time": "17:00",
            "end_time": "22:00",
            "rate_per_hour": 1.5
          }
        ],
        "flat_entry_fee": {
          "start_time": "22:00",
          "end_time": "07:30",
          "fee": 5.0
        },
        "first_hour_rate": 3.0,
        "max_daily_fee": 20.0
      },
      "saturday": {
        "time_based": [
          {
            "start_time": "07:30",
            "end_time": "17:00",
            "rate_per_hour": 2.5
          },
          {
            "start_time": "17:00",
            "end_time": "22:00",
            "rate_per_hour": 2.0
          }
        ],
        "flat_entry_fee": null,
        "first_hour_rate": 4.0,
        "max_daily_fee": 20.0
      },
      "sunday_public_holiday": {
        "time_based": [
          {
            "start_time": "07:30",
            "end_time": "17:00",
            "rate_per_hour": 3.0
          },
          {
            "start_time": "17:00",
            "end_time": "22:00",
            "rate_per_hour": 2.5
          }
        ],
        "flat_entry_fee": {
          "start_time": "22:00",
          "end_time": "07:30",
          "fee": 6.0
        },
        "first_hour_rate": null,
        "max_daily_fee": 20.0
      }
    }
  }
'''

    systemPrompt = f'''
    You are a model that excels at mapping out the pricing of carpark rates in Singapore. Given the dataset of carpark rates in Singapore, output a JSON string that contains the pricing of each carpark in the given dataset in a structured format. 

    This is the given template for the output:
    {schema}
    
    You are to output a list of JSON objects that follow this template, where
    - carpark_name is the name of the carpark
    - rates is an object that contains the pricing information for the carpark
    - weekday is an object that contains the pricing information for weekdays
    - saturday is an object that contains the pricing information for Saturdays
    - sunday_public_holiday is an object that contains the pricing information for Sundays and Public Holidays
    - time_based is an array of objects that contain the pricing information for different time slots
    - flat_entry_fee is an object that contains the pricing information for a flat entry fee
    - first_hour_rate is an object that contains the pricing information for the first hour rate
    - max_daily_fee is the maximum daily fee that can be charged for parking in the carpark
    - start_time and end_time are the start and end times for the time slot
    - rate_per_hour is the rate per hour for parking in the carpark
    - fee is the entry fee for parking in the carpark
    - rate is the rate for the first hour of parking in the carpark
    - subsequent_rate is the rate for subsequent hours of parking in the carpark

    If there are any fields where you are unable to determine the pricing, you should output a placeholder value of None.

    It doesn't matter if all the information in the dataset is not used, as long as the output is in the correct format, and the most accurate pricing information is displayed in the output.

    If you see the term or phrasing such as "Cap at $5.35" as an example, it means the fee is capped at $5.35. You should output the capped fee as the maximum daily fee in the output JSON string. Do NOT put this into the rate per hour field, nor the entry fee field.

    There can be instances where carparks do not charge by rate per hour but instead by entry fee. In such cases, you should output the entry fee as the flat entry fee in the output JSON string and leave the rate per hour empty. If there is no information available for rate during that time slot but an entry fee, you should assume that the cost of parking is the entry fee.

    If there are timeslots or periods of time where no information on the rate per hour or the flat entry fee that can be found, then do NOT create a timeslot for that period of time. You should only output the timeslots where there is information available.
    '''

    userPrompt = f'''
    Given this dataset of carpark rates in Singapore, output a JSON string that contains the pricing of each carpark in the given dataset in a structured format.
    {carpark_rates}


    The value for entryFee and the rates should be in dollars and be in the exact format as you see in the given template. Additionally, the fee should be "$0.00" if the fee is free or costs nothing. If there is no information available for a particular field, you should output a placeholder value of None.
    '''

    response = ask_gpt_json(systemPrompt, userPrompt)
    
    print(response)

    return response

# Write the entity's attributes into a row in a new excel sheet 
def write_entity_to_excel(entity_dict, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    if sheet.max_row == 1:
        headers = list(entity_dict.keys())
        sheet.append(headers)

    # Write values
    values = entity_dict.values()
    sheet.append(values)

    workbook.save(file_path)

def main():
    # Change this variable depending on if you're using gpt to format the data or calling it straight from the excel file.
    use_gpt = False

    # Initialize entity_dict to create later
    entity_dict = {}
    # id 
    # type
    # CarparkName
    # location
    # coordinates
    # ParkingCapacity
    # Sheltered
    # ParkingAvailability
    # Pricing   

    # Fetch carpark availabilities
    carpark_availabilities = fetch_carpark_availabilities()

    # Fetch carpark rates
    if use_gpt == True:
        try:
            # Fetch carpark rates
            carpark_rates = read_excel('CommercialCarparkRates.xlsx')
            formatted_carpark_rates = format_carpark_rates(carpark_rates)
        except:
            print("Error formatting carpark rates using GPT-3, check if the API key is correct.")
    else:
        try:
            # Open the Excel file
            wb = openpyxl.load_workbook('formatted_carpark_rates.xlsx')
            ws = wb.active

            # Initialize the carpark rates dictionary
            carpark_rates = {}

            # Iterate through the rows of the worksheet starting from the second row
            for row in ws.iter_rows(min_row=2, values_only=True):
                location, day_type, start_time, end_time, rate_per_hour, flat_entry_start, flat_entry_end, flat_entry_fee, first_hour_rate, max_daily_fee = row
                
                # Initialize carpark if not already in dict
                if location not in carpark_rates:
                    carpark_rates[location] = {'rates': {}}
                
                # Initialize day_type if not already in dict
                if day_type not in carpark_rates[location]['rates']:
                    carpark_rates[location]['rates'][day_type] = {
                        'time_based': [],
                        'flat_entry_fee': None,
                        'first_hour_rate': None,
                        'max_daily_fee': None
                    }
                
                # Add time-based rates
                if start_time != '-' and end_time != '-' and rate_per_hour != '-':
                    carpark_rates[location]['rates'][day_type]['time_based'].append({
                        'start_time': start_time,
                        'end_time': end_time,
                        'rate_per_hour': float(rate_per_hour)
                    })
                
                # Add flat entry fee (if exists)
                if flat_entry_start != '-' and flat_entry_end != '-' and flat_entry_fee != '-':
                    carpark_rates[location]['rates'][day_type]['flat_entry_fee'] = {
                        'start_time': flat_entry_start,
                        'end_time': flat_entry_end,
                        'fee': float(flat_entry_fee)
                    }
                else:
                    carpark_rates[location]['rates'][day_type]['flat_entry_fee'] = None
                
                # Add first hour rate
                if first_hour_rate != '-':
                    carpark_rates[location]['rates'][day_type]['first_hour_rate'] = float(first_hour_rate)
                else:
                    carpark_rates[location]['rates'][day_type]['first_hour_rate'] = None
                
                # Add max daily fee
                if max_daily_fee != '-':
                    carpark_rates[location]['rates'][day_type]['max_daily_fee'] = max_daily_fee
                else:
                    carpark_rates[location]['rates'][day_type]['max_daily_fee'] = None
                
            formatted_carpark_rates = carpark_rates
    
        except:
            print("Error reading from Excel file, check if the name of the file is correct, or if it's in the right directory.")

    # print("Formatted carpark rates:")  
    # print("=====================================")  
    # print(formatted_carpark_rates)
    # print("=====================================")

    # print("Carpark availabilities:")
    # print("=====================================")
    # print(carpark_availabilities)
    # print("=====================================")
    
    for key, value in carpark_availabilities.items():
        # initialize entity
        # e_id is the concatenation of the Development and CarParkID
        e_id = value['Development'].replace(" ", "") + str(value['CarParkID'])
        entity = Entity("Carpark", e_id, ctx=ctx)

        # set properties

        # CarparkName
        entity.prop("CarparkName", value['Development'])

        # Location
        coordinates = value['Location'].split(" ")
        entity.gprop("location", (float(coordinates[0]), float(coordinates[1])))

        # ParkingCapacity
        entity.prop("ParkingCapacity", "-")

        # Sheltered
        entity.prop("Sheltered", False)

        # ParkingAvailability
        entity.prop("ParkingAvailability", value['AvailableLots'])

        # Pricing
        # Fetch the carpark rates
        for carpark_name, carpark_data in formatted_carpark_rates.items():
            if carpark_name.strip().lower() == value['Development'].strip().lower():
                # print(f"CA_carpark_name: {value['Development']}")
                # print(f"CR_carpark_name: {carpark_name}")
                pricing = carpark_data
                entity.prop('Pricing', pricing)
                if carpark_name == "Bugis+":
                    print(carpark_data)
            
        entity_dict[e_id] = entity

    try:
        create_entities_in_broker(entity_dict.values())
    except:
        print("Failed to create entities, check if your context broker is running in Docker.")

    return entity_dict

# TEST CODE FOR SAVING LLM OUTPUT TO EXCEL
# =============================================================================
# def generate_formatted_excel(file_path):
    # carpark_rates = read_excel("CommercialCarparkRates.xlsx")
    # response = format_carpark_rates(carpark_rates)
    # print(type(response))

    # if os.path.exists(file_path):
    #     workbook = openpyxl.load_workbook(file_path)
    #     sheet = workbook.active
    # else:
    #     workbook = openpyxl.Workbook()
    #     sheet = workbook.active

    # headers = ["Location", "Day", "Start Time", "End Time", "Rate Per Hour", "Flat Entry Start", "Flat Entry End", "Flat Entry Fee", "First Hour Rate", "Max Daily Fee"]
    # sheet.append(headers)

    # for location, carpark_data in response.items():
    #     for day, rates in carpark_data['rates'].items():
    #         time_based = rates.get('time_based', '-')
    #         flat_entry_fee = rates.get('flat_entry_fee')
    #         if isinstance(flat_entry_fee, dict):
    #             flat_start = flat_entry_fee['start_time']
    #             flat_end = flat_entry_fee['end_time']
    #             flat_fee = flat_entry_fee['fee']
    #         else:
    #             flat_start = flat_end = flat_fee = '-'
    #         first_hour_rate = rates.get('first_hour_rate', '-')
    #         max_daily_fee = rates.get('max_daily_fee', '-')
            
    #         # If there is a time-based rate, write each period
    #         if time_based and isinstance(time_based, list):
    #             for tb in time_based:
    #                 sheet.append([
    #                     location, day, 
    #                     tb.get('start_time', '-'), tb.get('end_time', '-'), tb.get('rate_per_hour', '-'),
    #                     flat_start, flat_end, flat_fee,
    #                     first_hour_rate if first_hour_rate is not None else '-',
    #                     max_daily_fee if max_daily_fee is not None else '-'
    #                 ])
    #         else:
    #             # If no time-based rate, write a single row with default values
    #             sheet.append([
    #                 location, day, '-', '-', '-', 
    #                 flat_start, flat_end, flat_fee,
    #                 first_hour_rate if first_hour_rate is not None else '-',
    #                 max_daily_fee if max_daily_fee is not None else '-'
    #             ])

    # workbook.save(file_path)
    # workbook.close()
# =============================================================================

# TEST MAIN()
# =============================================================================
print(main())
# =============================================================================