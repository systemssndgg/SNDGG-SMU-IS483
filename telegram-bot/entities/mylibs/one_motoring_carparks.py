from . import constants
from landtransportsg import Traffic
import requests
from openai import OpenAI
import openpyxl
import os
import re
import googlemaps
import time

import json
from requests.exceptions import RequestException, HTTPError
from ngsildclient import Client, Entity, SmartDataModels
from datetime import datetime
import pandas as pd

# client = OpenAI(
#     api_key = constants.OPENAPI_KEY
# )

API_KEY = constants.LTA_API_KEY
ctx = constants.ctx
broker_url = constants.broker_url
broker_port = constants.broker_port  # default, 80
temporal_port = constants.temporal_port  # default 1026
broker_tenant = constants.broker_tenant


# (1) Format the carpark rates using GPT-4o
# =============================================================================
# In order to avoid repeated calls to GPT, we can save the formatted data to an Excel file and read from it.
# For now, ensure that you have one_motoring.xlsx, and FormattedOneMotoring.xlsx in the same directory as this script.
# one_motoring.xlsx is the raw data of carpark rates from OneMotoring.
# FormattedOneMotoring.xlsx is the formatted version of one_motoring.xlsx using GPT.
# =============================================================================

# HELPER FUNCTIONS
# =============================================================================
# This function helps to generate the formatted excel file using GPT and CommercialCarparkRates.xlsx
def generate_formatted_excel(new_file_path):
    '''
    Inputs:
        file_path: Name of the new excel sheet created.
    Outputs:
        None
    Description:
        Saves the formatted output from GPT into an excel sheet that is named after file_path.
    '''

    file_path = os.path.join(os.path.dirname(__file__), 'one_motoring.xlsx')
    carpark_rates = read_excel(file_path) 
    # There is a total of 412 carparks in the dataset of one_motoring.xlsx
    # Need to split the the carparks into batches of 10 to avoid hitting the GPT-4o token limit
    # Split the carpark rates into batches of 10
    carpark_rates_batches = [carpark_rates[i:i + 10] for i in range(0, len(carpark_rates), 10)]
    # For each batch, format the carpark rates using GPT-4o, and save it to the new excel file or append it if the file already exists.
    for batch in carpark_rates_batches:        
        count = 1
        print("Formatting carpark rates using GPT-4o...")
        print("Batch: ", count)
        response = format_carpark_rates(batch)
        print("Done formatting Batch: ", count)
        count += 1

        if os.path.exists(new_file_path):
            workbook = openpyxl.load_workbook(new_file_path)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

        if sheet.max_row == 1:
            headers = ["Location", "Day", "Start Time", "End Time", "Rate Per Hour", "Flat Entry Start", "Flat Entry End", "Flat Entry Fee", "First Hour Rate", "Max Daily Fee"]
            sheet.append(headers)

        for location, carpark_data in response.items():
            for day, rates in carpark_data['rates'].items():
                time_based = rates.get('time_based', '-')
                flat_entry_fee = rates.get('flat_entry_fee')
                if isinstance(flat_entry_fee, dict):
                    flat_start = flat_entry_fee['start_time']
                    flat_end = flat_entry_fee['end_time']
                    flat_fee = flat_entry_fee['fee']
                else:
                    flat_start = flat_end = flat_fee = '-'
                first_hour_rate = rates.get('first_hour_rate', '-')
                max_daily_fee = rates.get('max_daily_fee', '-')
                
                # If there is a time-based rate, write each period
                if time_based and isinstance(time_based, list):
                    for tb in time_based:
                        sheet.append([
                            location, day, 
                            tb.get('start_time', '-'), tb.get('end_time', '-'), tb.get('rate_per_hour', '-'),
                            flat_start, flat_end, flat_fee,
                            first_hour_rate if first_hour_rate is not None else '-',
                            max_daily_fee if max_daily_fee is not None else '-'
                        ])
                else:
                    # If no time-based rate, write a single row with default values
                    sheet.append([
                        location, day, '-', '-', '-', 
                        flat_start, flat_end, flat_fee,
                        first_hour_rate if first_hour_rate is not None else '-',
                        max_daily_fee if max_daily_fee is not None else '-'
                    ])

        print("Saving formatted data to Excel file...")	
        new_file_path = os.path.join(os.path.dirname(__file__), new_file_path)
        workbook.save(new_file_path)
        workbook.close()
        
# This function invokes Geocoding API from Google to return the coordinates of each carpark in one_motoring.xlsx
def find_coordinates(location_name, api_key):
    """
    Inputs: 
        location_name : str : Name of the location
        api_key : str : Your Google Maps API key
    Outputs:
        tuple : A tuple of (latitude, longitude) or None if not found
    Description: 
        Given the name of a location, return the latitude and longitude.
    """
    # Initialize the Google Maps client
    gmaps = googlemaps.Client(key=api_key)
    
    # Perform geocoding
    geocode_result = gmaps.geocode(location_name)
    
    if geocode_result:
        # Extract latitude and longitude
        location = geocode_result[0]["geometry"]["location"]
        return location["lat"], location["lng"]
    else:
        print("Location not found.")
        return None

# This is a helper function to make a call to GPT-4o-mini
def ask_gpt_json(system_prompt, user_prompt):
        '''
            Input:
                system_prompt: str : The system prompt to provide to the model 
                user_prompt: str : The user prompt to provide to the model 

            Output:
                response: dict : The response generated by the model 
                dict structure based on the examples given in system prompt.
                cost: float : The cost of the API call
        '''

        model="gpt-4.o",
        completion = client.chat.completions.create(
            # Adjust the temperature to control the randomness of the output
            temperature=0.2,
            model="gpt-4o-mini",
            # Output will be a string in valid JSON format, so we can parse it with json.loads
            response_format= {"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": system_prompt
                },
                {
                    "role": "user",
                    "content": user_prompt
                }
            ]
        )
        response = completion.choices[0].message.content
        try:
            f = open("response.txt", "w")
            f.write(response)
            f.close()
            print("Data formatted and written to file")
        except:
                print("Error writing data to file")

        # Converts the response to a JSON Object / Dictionary.
        response = json.loads(response)
        return response

# This function formats the carpark rates using GPT-4o
def format_carpark_rates(carpark_rates):
    schema = '''
{
  "Carpark A": {
    "rates": {
      "weekday": {
        "time_based": [
          {
            "start_time": "07:30",
            "end_time": "17:00",
            "rate_per_hour": 2.0
          },
          {
            "start_time": "17:00",
            "end_time": "22:00",
            "rate_per_hour": 1.5
          }
        ],
        "flat_entry_fee": {
          "start_time": "22:00",
          "end_time": "07:30",
          "fee": 5.0
        },
        "first_hour_rate": 3.0,
        "max_daily_fee": 20.0
      },
      "saturday": {
        "time_based": [
          {
            "start_time": "07:30",
            "end_time": "17:00",
            "rate_per_hour": 2.5
          },
          {
            "start_time": "17:00",
            "end_time": "22:00",
            "rate_per_hour": 2.0
          }
        ],
        "flat_entry_fee": null,
        "first_hour_rate": 4.0,
        "max_daily_fee": 20.0
      },
      "sunday_public_holiday": {
        "time_based": [
          {
            "start_time": "07:30",
            "end_time": "17:00",
            "rate_per_hour": 3.0
          },
          {
            "start_time": "17:00",
            "end_time": "22:00",
            "rate_per_hour": 2.5
          }
        ],
        "flat_entry_fee": {
          "start_time": "22:00",
          "end_time": "07:30",
          "fee": 6.0
        },
        "first_hour_rate": null,
        "max_daily_fee": 20.0
      }
    }
  }
'''

    systemPrompt = f'''
    You are a model that excels at mapping out the pricing of carpark rates in Singapore. Given the dataset of carpark rates in Singapore, output a JSON string that contains the pricing of each carpark in the given dataset in a structured format. 

    This is the given template for the output:
    {schema}
    
    You are to output a list of JSON objects that follow this template, where
    - carpark_name is the name of the carpark
    - rates is an object that contains the pricing information for the carpark
    - weekday is an object that contains the pricing information for weekdays
    - saturday is an object that contains the pricing information for Saturdays
    - sunday_public_holiday is an object that contains the pricing information for Sundays and Public Holidays
    - time_based is an array of objects that contain the pricing information for different time slots
    - flat_entry_fee is an object that contains the pricing information for a flat entry fee
    - first_hour_rate is an object that contains the pricing information for the first hour rate
    - max_daily_fee is the maximum daily fee that can be charged for parking in the carpark
    - start_time and end_time are the start and end times for the time slot
    - rate_per_hour is the rate per hour for parking in the carpark
    - fee is the entry fee for parking in the carpark
    - rate is the rate for the first hour of parking in the carpark
    - subsequent_rate is the rate for subsequent hours of parking in the carpark

    If there are any fields where you are unable to determine the pricing, you should output a placeholder value of None.

    It doesn't matter if all the information in the dataset is not used, as long as the output is in the correct format, and the most accurate pricing information is displayed in the output.

    If you see the term or phrasing such as "Cap at $5.35" as an example, it means the fee is capped at $5.35. You should output the capped fee as the maximum daily fee in the output JSON string. Do NOT put this into the rate per hour field, nor the entry fee field.

    There can be instances where carparks do not charge by rate per hour but instead by entry fee. In such cases, you should output the entry fee as the flat entry fee in the output JSON string and leave the rate per hour empty. If there is no information available for rate during that time slot but an entry fee, you should assume that the cost of parking is the entry fee.

    If there are timeslots or periods of time where no information on the rate per hour or the flat entry fee that can be found, then do NOT create a timeslot for that period of time. You should only output the timeslots where there is information available.
    '''

    userPrompt = f'''
    Given this dataset of carpark rates in Singapore, output a JSON string that contains the pricing of each carpark in the given dataset in a structured format.
    {carpark_rates}


    The value for entryFee and the rates should be in dollars and be in the exact format as you see in the given template. Additionally, the fee should be "$0.00" if the fee is free or costs nothing. If there is no information available for a particular field, you should output a placeholder value of None.
    '''

    response = ask_gpt_json(systemPrompt, userPrompt)
    
    print(response)

    return response

# This creates entities in the broker
def create_entities_in_broker(entities):
    with Client(
        hostname=broker_url,
        port=broker_port,
        tenant=broker_tenant,
        port_temporal=temporal_port,
    ) as client:
        count = 0
        for entity in entities:
            ret = client.upsert(entity)
            if ret:
                count += 1
    print("Uploaded ", count)
    return ret

# This function writes the entity to an excel file
def write_entity_to_excel(entity_dict, file_path):
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        headers = list(entity_dict.keys())
        sheet.append(headers)

    # Write values
    values = list(entity_dict.values())
    sheet.append(values)

    workbook.save(file_path)

# This is a helper function to read the excel files
def read_excel(file_path):
    '''
    Inputs: file_path : str : The path to the excel file to read
    \nOutputs: data : list : A list of lists, where each list corresponds to a row, containing the data from the excel file

    \nDescription: Reads the excel file at the given file path and returns the data in the form of a list of lists.
    '''
    file_path = os.path.join(os.path.dirname(__file__), file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = [row for row in sheet.iter_rows(values_only=True)]
    return data

# Helper function to load the raw carpark rates from one_motoring.xlsx
def fetch_carpark_rates(carpark_name, file_path='one_motoring.xlsx'): #'entities/mylibs/one_motoring.xlsx'
    file_path = os.path.join(os.path.dirname(__file__), file_path)
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook.active

    # Find the headers and get column indices
    headers = [cell.value for cell in sheet[1]]
    carpark_col = headers.index('carpark')
    weekdays_rate_1_col = headers.index('weekdays_rate_1')
    weekdays_rate_2_col = headers.index('weekdays_rate_2')
    saturday_rate_col = headers.index('saturday_rate')
    sunday_rate_col = headers.index('sunday_publicholiday_rate')

    # Initialize variables to hold the rates
    weekday_rate_combined, saturday_rate, sunday_rate = None, None, None

    # Iterate over the rows to find the specific carpark
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[carpark_col] == carpark_name:
            # Combine weekday rates
            weekday_rate_1 = row[weekdays_rate_1_col] or ""
            weekday_rate_2 = row[weekdays_rate_2_col] or ""
            weekday_rate_combined = f"{weekday_rate_1}; {weekday_rate_2}".strip("; ")

            # Extract Saturday and Sunday rates
            saturday_rate = row[saturday_rate_col]
            sunday_rate = row[sunday_rate_col]

            break  # Exit the loop once the carpark is found

    # Return the pricing as a dictionary
    return {
        'WeekdayStr': weekday_rate_combined or "-",
        'SaturdayStr': saturday_rate or "-",
        'SundayPHStr': sunday_rate or "-"
    }

# Helper function to create a carpark entity
def create_entity(carpark_name="-", e_id="-", coordinates="-", sheltered="-", parking_capacity="-", available_lots="-", pricing="-"):
    '''
    Inputs:
        carpark_name : str : The name of the carpark
        e_id : str : The entity ID of the carpark
        coordinates : tuple : The coordinates of the carpark in the format (latitude, longitude)
        sheltered : bool : Whether the carpark is sheltered or not
        parking_capacity : int : The total number of parking lots in the carpark
        available_lots : int : The number of available parking lots in the carpark
        pricing : dict : The pricing information for the carpark
    Outputs:
        entity : Entity : The entity object created for the carpark
        
    Description:
        This function creates an entity object for a carpark with the given parameters. If any of the parameters are missing, the key will not exist.
        '''
    entity = Entity("Carpark", e_id, ctx=ctx)

    # Set properties
    if carpark_name != "-":
        entity.prop("carparkName", carpark_name)
    if coordinates != "-" and coordinates != None:
        entity.gprop("location", (float(coordinates[0]), float(coordinates[1])))
    if (sheltered != "-"):
        entity.prop("sheltered", True) # Assume all the carparks are sheltered
    if (parking_capacity != "-"):
        entity.prop("parkingCapacity", parking_capacity)
    if (available_lots != "-"):
        entity.prop("parkingAvailability", available_lots)
    if pricing != "-":
        entity.prop('pricing', pricing)

    return entity

def clean_text(text):
    return re.sub(r'[^\x00-\x7F]+', ' ', text)  # Removes non-ASCII characters

# This function ties in everything to create the entities and push it to the context broker.
def create_one_motoring_carparks():
    # Change this variable depending on if you're using gpt to format the data or calling it straight from the excel file.
    use_gpt = False

    # Initialize entity_dict to store entities
    entity_dict = {}
    # id 
    # type
    # CarparkName
    # location
    # coordinates
    # ParkingCapacity
    # Sheltered
    # ParkingAvailability
    # Pricing   

    # Fetch carpark rates
    if use_gpt == True:
        try:
            # Fetch carpark rates
            generate_formatted_excel("FormattedOneMotoring.xlsx")
            file_path = os.path.join(os.path.dirname(__file__), 'FormattedOneMotoring.xlsx')
            carpark_rates = read_excel(file_path)
            formatted_carpark_rates = format_carpark_rates(carpark_rates)
        except:
            print("Error formatting carpark rates using GPT-4o, check if the API key is correct.")
    else:
        try:
            # Open the Excel file
            # file_path = os.path.join(os.path.dirname(__file__), 'FormattedOneMotoring.xlsx')
            wb = openpyxl.load_workbook("entities/mylibs/FormattedOneMotoring.xlsx")
            ws = wb.active

            # Initialize the carpark rates dictionary
            carpark_rates = {}

            # Iterate through the rows of the worksheet starting from the second row
            for row in ws.iter_rows(min_row=2, values_only=True):
                location, day_type, start_time, end_time, rate_per_hour, flat_entry_start, flat_entry_end, flat_entry_fee, first_hour_rate, max_daily_fee = row
                
                # Initialize carpark if not already in dict
                if location not in carpark_rates:
                    carpark_rates[location] = {'rates': {}}
                
                # Initialize day_type if not already in dict
                if day_type not in carpark_rates[location]['rates']:
                    carpark_rates[location]['rates'][day_type] = {
                        'timeBased': []
                    }
                
                # Add time-based rates
                if start_time != '-' and end_time != '-' and rate_per_hour != '-':
                    carpark_rates[location]['rates'][day_type]['timeBased'].append({
                        'startTime': start_time,
                        'endTime': end_time,
                        'ratePerHour': float(rate_per_hour)
                    })
                
                # Add flat entry fee (if exists)
                if flat_entry_start != '-' and flat_entry_end != '-' and flat_entry_fee != '-':
                    carpark_rates[location]['rates'][day_type]['flatEntryFee'] = {
                        'startTime': flat_entry_start,
                        'endTime': flat_entry_end,
                        'fee': float(flat_entry_fee)
                    }
                
                # Add first hour rate
                if first_hour_rate != '-':
                    carpark_rates[location]['rates'][day_type]['firstHourRate'] = float(first_hour_rate)
                
                # Add max daily fee
                if max_daily_fee != '-':
                    carpark_rates[location]['rates'][day_type]['maxDailyFee'] = max_daily_fee
                
            formatted_carpark_rates = carpark_rates
    
        except:
            print("Error reading from Excel file, check if the name of the file is correct, or if it's in the right directory.")

    # General case for other carparks
    id = 1
    for carpark_name, carpark_data in formatted_carpark_rates.items():
        # Create the entity ID (e.g., OM1, OM2, ...)
        e_id = "OM" + str(id)
        id += 1

        # Fetch carpark coordinates using Geocoding API
        # Coordinates have been previously fetched and stored in om_carpark_coordinates.xlsx
        coordinates = "-"
        # coord_path = os.path.join(os.path.dirname(__file__), 'om_carpark_coordinates.xlsx')
        workbook = openpyxl.load_workbook('entities/mylibs/om_carpark_coordinates.xlsx')
        sheet = workbook.active
        data = [row for row in sheet.iter_rows(values_only=True)]
        # carpark name is the first column, latitude is the second column, longitude is the third column
        # check for the corresponding carpark name, and get the coordinates
        for row in data:
            if row[0] == carpark_name:
                coordinates = (row[1], row[2])
                break
        if coordinates == "-":
            # Call the Geocoding API to get the coordinates
            GOOGLE_API_KEY = constants.GOOGLE_MAPS_KEY
            location_name = carpark_name + ", Singapore"
            try:
                coordinates = find_coordinates(location_name, GOOGLE_API_KEY)
                if coordinates:
                    print(f"The coordinates of {location_name} are: {coordinates}")
                else:
                    print("Could not find coordinates for the location.")
                    coordinates = "-"
            except googlemaps.exceptions.ApiError as e:
                print(f"API error occurred: {e}")
                coordinates = "-"
            except googlemaps.exceptions.TransportError as e:
                print(f"Network error occurred: {e}")
                coordinates = "-"
            except Exception as e:
                print(f"An error occurred: {e}")
                coordinates = "-"
        
        # No availability can be discerned

        # Assume all of them are sheltered
        sheltered = True

        # Add the formatted rates to the pricing dictionary
        pricing = {}
        pricing['rates'] = carpark_data['rates']
        raw_rates = fetch_carpark_rates(carpark_name)
        pricing['weekdayStr'] = clean_text(raw_rates['WeekdayStr'])
        pricing['saturdayStr'] = clean_text(raw_rates['SaturdayStr'])
        pricing['sundayPHStr'] = clean_text(raw_rates['SundayPHStr'])

        # Create and store the entity
        entity = create_entity(carpark_name=carpark_name, e_id=e_id, sheltered=sheltered, coordinates=coordinates, pricing=pricing)
        entity_dict[e_id] = entity
    
    # try:
    #     create_entities_in_broker(entity_dict.values())
    # except:
    #     print("Failed to create entities, check if your context broker is running in Docker.")

    return list(entity_dict.values())


